import openpyxl
from django.http import HttpResponse
from api.models import NewRobot
from django.core import serializers
import json
from openpyxl import Workbook


def excel_report_view(request):
    # new_robots_data = serializers.serialize('json', NewRobot.objects.all())
    # new_robots_data = json.loads(new_robots_data)
    #
    # wb = Workbook()
    # ws = wb.active
    # ws.title = 'Robots'
    # headers = ['serial', 'model', 'version']
    # ws.append(headers)
    # for data in new_robots_data:
    #     serial = data['fields']['serial']
    #     model = data['fields']['model']
    #     version = data['fields']['version']
    #     ws.append([serial, model, version])
    #
    # response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="excel_report.xlsx"'
    # wb.save(response)
    # return response
    data = NewRobot.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Robots data'

    sheet['A1'] = 'Модель'
    sheet['B1'] = 'Версия'
    sheet['C1'] = 'Количество за неделю'

    row_number = 2
    for row in data:
        sheet.cell(row=row_number, column=1).value = row.model
        sheet.cell(row=row_number, column=2).value = row.version
        robots_count = NewRobot.objects.filter(model=row.model).filter(version=row.version).all()
        sheet.cell(row=row_number, column=3).value = len(robots_count)
        row_number += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robots_data.xlsx'
    workbook.save(response)
    return response
